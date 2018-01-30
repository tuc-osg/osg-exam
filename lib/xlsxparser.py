from openpyxl import load_workbook


def students(xlsx_file):
    wb = load_workbook(filename=xlsx_file, read_only=True)
    sheets = wb.get_sheet_names()
    ws = wb[sheets[0]]

    startRow = 1
    matrikelCol = 1
    nameCol = 1
    vornameCol = 1

    matrikelFound = False
    nameFound = False
    vornameFound = False

    # find the starting row for the data and the columns for id, name and first name
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if str(ws.cell(row=i, column=j).value).lower() in ['matrikelnummer', 'matr-nr.']:
                startRow = i + 1
                matrikelCol = j
                matrikelFound = True
            elif 'name' == str(ws.cell(row=i, column=j).value).lower():
                nameCol = j
                nameFound = True
            elif 'vorname' == str(ws.cell(row=i, column=j).value).lower():
                vornameCol = j
                vornameFound = True

    if not (matrikelFound and nameFound and vornameFound):
        raise RuntimeError("Could not find neccessary columns in {0}.".format(xlsx_file))

    for i in range(startRow, ws.max_row + 1):
        student_id = str(ws.cell(row=i, column=matrikelCol).value)
        first_name = ws.cell(row=i, column=vornameCol).value
        last_name = ws.cell(row=i, column=nameCol).value
        if student_id != 'None':  # Might be None in empty rows
            yield first_name, last_name, student_id
