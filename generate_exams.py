'''
	Script to generate personalized student exams, and a blank one.

	Assumptions:
	- The script can re-generate student.tex and compile exam.tex to get a PDF.
	- The file teilnehmer.xlsx exists and has the following format:
			unchanged(default) worksheet name: Table 1
			Columns with "Matrikelnummer", "Vorname" and "Name"
			...
'''
import os, shutil, os.path
from openpyxl import Workbook
from openpyxl import load_workbook

# Target folder for exam PDF files.
folder="pdf"

def genexam(lastname, firstname, studentid):
	'''
		Generates an exam for the given student, if not existing.
		Target file is stored in "folder", with the student ID as file name.
	'''
	targetfile = folder+os.sep+studentid+'.pdf'
	if os.path.isfile(targetfile):
		print("Skipping generation of %s, because it exists ..."%targetfile)
		return
	else:
		print("Generating "+targetfile)
	studentfile=open('student.tex','w')
	studentfile.write('''
\\newcommand{\\studentid}{%s}\n
\\newcommand{\\studentname}{%s %s}
	'''%(studentid, firstname, lastname))
	studentfile.close()
	os.system('latexmk -C')
	os.system('lualatex exam.tex')  # double run needed by exam class
	os.system('lualatex exam.tex')
	shutil.move('exam.pdf', targetfile)

try:
	os.mkdir(folder)
except:
	pass


# with open('teilnehmer.csv', 'rb') as csvfile:
# 	reader = csv.reader(csvfile, delimiter=';')
# 	for row in reader:
# 		genexam(*row)
#
# # Also generate a blank one. LaTex needs some content, so we use gibberish white space.
# genexam('~','~','~')

wb = load_workbook(filename='zulassungsliste.xlsx', read_only=True)
sheets = wb.get_sheet_names()
ws = wb[sheets[0]]

startRow = 1
matrikelCol = 1
# nameRow = 1
nameCol = 1
# vornameRow = 1
vornameCol = 1

for i in range(1,ws.max_row+1):
	for j in range(1, ws.max_column+1):
		if 'matrikelnummer' == str(ws.cell(row=i,column=j).value).lower():
			startRow = i
			matrikelCol = j
		elif 'name' == str(ws.cell(row=i, column=j).value).lower():
			# nameRow = i
			nameCol = j
		elif 'vorname' == str(ws.cell(row=i, column=j).value).lower():
			# vornameRow = i
			vornameCol = j
		# print(ws.cell(row=i,column=j).value)


for i in range(startRow,ws.max_row+1):
	# student = (ws.cell(row=i,column=matrikelCol).value, ws.cell(row=i,column=vornameCol).value, ws.cell(row=i,column=nameCol).value)
	# print(student)
	genexam(ws.cell(row=i, column=matrikelCol).value, ws.cell(row=i, column=vornameCol).value, ws.cell(row=i, column=nameCol).value)
