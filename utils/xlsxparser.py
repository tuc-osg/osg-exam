from openpyxl import load_workbook


def students(xlsx_file):
    wb = load_workbook(filename=xlsx_file, read_only=True)
    sheets = wb.get_sheet_names()
    ws = wb[sheets[0]]

    # find the starting row for the data and the columns for id, name and first name
    student_id_col = None
    first_name_col = None
    last_name_col = None
    for row_index, row in enumerate(ws.iter_rows()):
        if student_id_col and first_name_col and last_name_col:
            student_id = str(ws.cell(row=row_index+1, column=student_id_col+1).value)
            first_name = ws.cell(row=row_index+1, column=first_name_col+1).value
            last_name = ws.cell(row=row_index+1, column=last_name_col+1).value
            if student_id != 'None':  # Might be None in empty rows
                yield first_name, last_name, student_id
        else:
            for col_index, cell in enumerate(row):
                if str(cell.value).lower() in ['matrikelnummer', 'matr-nr.']:
                    student_id_col = col_index
                elif 'name' == str(cell.value).lower():
                    last_name_col = col_index
                elif 'vorname' == str(cell.value).lower():
                    first_name_col = col_index
